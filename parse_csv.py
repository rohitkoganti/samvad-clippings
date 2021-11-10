import openpyxl


def parse_csv(filename):
    headlines = []

    #Open csv file
    #Parse the contents for headlines, source, journalist, link
    #Push these as a tuple into content

    wb= openpyxl.load_workbook(filename)
    sheet = wb.active
    print("Parse: Attempting to parse headlines...")

    if sheet['A3'].value == 'S. NO.' and sheet['D3'].value == 'CONTENT' and sheet['E3'].value == 'PUBLICATION' \
        and sheet['F3'].value == 'JOURNALIST' and sheet['G3'].value == 'NEWS LINK':
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
            headline = []
            for cell in row[:6]:
                headline.append(cell.value)
            headline.append(row[6].hyperlink.target)
            headlines.append(headline)
        print(headlines)
        print("Parse: Successfully parsed headlines.")
        return headlines
    else:
        print("Parse: Failed to read excel sheet, format might have changed")
        return None
